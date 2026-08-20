"""Tests for Excel report constructor spec validation and workbook assembly."""

from __future__ import annotations

import unittest
from datetime import datetime
from io import BytesIO

from openpyxl import load_workbook

from app.crm.reports.catalog import (
    CLOSURE_KIND_LABELS,
    DATASETS,
    NEST_NESTED,
    ReportSheetSpec,
    ReportSpec,
    catalog_payload,
    format_cell_value,
    preset_closed_orders_with_tasks,
    preset_surveyed_order_summary,
    validate_report_spec,
)
from app.crm.reports.errors import ReportError
from app.crm.reports.excel import build_workbook_bytes, content_disposition
from app.crm.reports.query import (
    SheetData,
    aggregate_surveyed_order_counts,
    resolve_surveyed_task_outcome,
)


class ValidateReportSpecTests(unittest.TestCase):
    def test_preset_is_valid(self) -> None:
        spec = validate_report_spec(preset_closed_orders_with_tasks())
        self.assertEqual(spec.sheets[0].dataset, "closed_orders")
        self.assertEqual(spec.sheets[1].parent_sheet, "orders")
        self.assertEqual(spec.sheets[1].dataset, "closed_tasks")

    def test_rejects_empty_sheets(self) -> None:
        with self.assertRaises(ReportError):
            validate_report_spec(ReportSpec(name="Пустой", sheets=[]))

    def test_rejects_unknown_dataset(self) -> None:
        with self.assertRaises(ReportError) as ctx:
            validate_report_spec(
                ReportSpec(
                    name="Bad",
                    sheets=[
                        ReportSheetSpec(id="a", dataset="no_such_dataset", title="X"),
                    ],
                )
            )
        self.assertIn("Неизвестный набор данных", str(ctx.exception))

    def test_rejects_unknown_column(self) -> None:
        with self.assertRaises(ReportError) as ctx:
            validate_report_spec(
                ReportSpec(
                    name="Bad",
                    sheets=[
                        ReportSheetSpec(
                            id="orders",
                            dataset="closed_orders",
                            columns=["not_a_column"],
                        )
                    ],
                )
            )
        self.assertIn("Неизвестные колонки", str(ctx.exception))

    def test_child_requires_parent(self) -> None:
        with self.assertRaises(ReportError):
            validate_report_spec(
                ReportSpec(
                    name="Bad",
                    sheets=[
                        ReportSheetSpec(id="tasks", dataset="closed_tasks", title="Задачи"),
                    ],
                )
            )

    def test_parent_must_exist(self) -> None:
        with self.assertRaises(ReportError):
            validate_report_spec(
                ReportSpec(
                    name="Bad",
                    sheets=[
                        ReportSheetSpec(
                            id="tasks",
                            dataset="closed_tasks",
                            parent_sheet="missing",
                        )
                    ],
                )
            )

    def test_rejects_wrong_parent_dataset(self) -> None:
        with self.assertRaises(ReportError):
            validate_report_spec(
                ReportSpec(
                    name="Bad",
                    sheets=[
                        ReportSheetSpec(id="field", dataset="field_summary"),
                        ReportSheetSpec(
                            id="tasks",
                            dataset="closed_tasks",
                            parent_sheet="field",
                        ),
                    ],
                )
            )

    def test_rejects_unknown_filter(self) -> None:
        with self.assertRaises(ReportError):
            validate_report_spec(
                ReportSpec(
                    name="Bad",
                    sheets=[
                        ReportSheetSpec(
                            id="orders",
                            dataset="closed_orders",
                            filters={"nope": ["x"]},
                        )
                    ],
                )
            )

    def test_empty_columns_means_all(self) -> None:
        spec = validate_report_spec(
            ReportSpec(
                name="Все колонки",
                sheets=[ReportSheetSpec(id="orders", dataset="closed_orders")],
            )
        )
        self.assertEqual(spec.sheets[0].columns, [])
        self.assertTrue(DATASETS["closed_orders"].columns)

    def test_catalog_lists_parent_child_links(self) -> None:
        payload = catalog_payload()
        datasets = {item["id"]: item for item in payload["datasets"]}
        self.assertIn("closed_orders", datasets)
        child_ids = {item["id"] for item in datasets["closed_orders"]["child_datasets"]}
        self.assertIn("closed_tasks", child_ids)
        self.assertIn("active_tasks_in_orders", child_ids)
        self.assertEqual(datasets["closed_tasks"]["parent_datasets"], ["closed_orders"])

    def test_surveyed_summary_preset_is_valid(self) -> None:
        spec = validate_report_spec(preset_surveyed_order_summary())
        self.assertEqual(len(spec.sheets), 1)
        self.assertEqual(spec.sheets[0].dataset, "surveyed_order_summary")
        self.assertIsNone(spec.sheets[0].parent_sheet)
        self.assertIn("pre_analise", spec.sheets[0].columns)
        self.assertIn("analise", spec.sheets[0].columns)

    def test_catalog_includes_surveyed_summary_and_clear_source(self) -> None:
        payload = catalog_payload()
        datasets = {item["id"]: item for item in payload["datasets"]}
        self.assertIn("surveyed_order_summary", datasets)
        column_ids = {col["id"] for col in datasets["surveyed_order_summary"]["columns"]}
        self.assertIn("pre_analise", column_ids)
        self.assertIn("analise", column_ids)
        source_values = {
            opt["value"]
            for flt in datasets["closed_tasks"]["filters"]
            if flt["id"] == "sources"
            for opt in flt["options"]
        }
        self.assertIn("clear", source_values)
        preset_ids = {item["id"] for item in payload["presets"]}
        self.assertEqual(
            preset_ids,
            {"closed_orders_with_tasks", "surveyed_order_summary"},
        )

    def test_surveyed_analise_flags_format_as_yes_no(self) -> None:
        dataset = DATASETS["surveyed_order_summary"]
        self.assertEqual(format_cell_value(dataset.column("pre_analise"), True), "Да")
        self.assertEqual(format_cell_value(dataset.column("analise"), False), "Нет")

    def test_closure_kind_clear_label(self) -> None:
        column = DATASETS["closed_tasks"].column("closure_kind")
        self.assertEqual(format_cell_value(column, "clear"), "Разрытие отсутствует")
        self.assertEqual(CLOSURE_KIND_LABELS["clear"], "Разрытие отсутствует")


class BuildWorkbookTests(unittest.TestCase):
    def test_flat_sheet_headers_and_values(self) -> None:
        dataset = DATASETS["closed_orders"]
        columns = [dataset.column("task_number"), dataset.column("rayon"), dataset.column("status")]
        sheets = [
            SheetData(
                id="orders",
                title="Заказы",
                dataset="closed_orders",
                columns=columns,
                rows=[
                    {
                        "task_number": "A-1",
                        "rayon": "Тверской",
                        "status": "done",
                    }
                ],
            )
        ]
        raw = build_workbook_bytes("Тест", sheets)
        wb = load_workbook(BytesIO(raw))
        ws = wb["Заказы"]
        self.assertEqual([cell.value for cell in ws[1]], ["Номер заказа", "Район", "Статус"])
        self.assertEqual(ws["A2"].value, "A-1")
        self.assertEqual(ws["B2"].value, "Тверской")
        self.assertEqual(ws["C2"].value, "Завершённые")
        self.assertIsNotNone(ws.auto_filter.ref)
        self.assertEqual(ws.freeze_panes, "A2")

    def test_surveyed_summary_headers(self) -> None:
        dataset = DATASETS["surveyed_order_summary"]
        columns = [
            dataset.column("task_number"),
            dataset.column("pre_analise"),
            dataset.column("analise"),
            dataset.column("tasks_surveyed"),
            dataset.column("tasks_clear"),
            dataset.column("tasks_done_legal"),
            dataset.column("tasks_done_illegal"),
            dataset.column("tasks_open"),
        ]
        sheets = [
            SheetData(
                id="orders",
                title="Заказы",
                dataset="surveyed_order_summary",
                columns=columns,
                rows=[
                    {
                        "task_number": "B-2",
                        "pre_analise": True,
                        "analise": False,
                        "tasks_surveyed": 4,
                        "tasks_clear": 1,
                        "tasks_done_legal": 1,
                        "tasks_done_illegal": 1,
                        "tasks_open": 1,
                    }
                ],
            )
        ]
        raw = build_workbook_bytes("Обследованные заказы", sheets)
        wb = load_workbook(BytesIO(raw))
        ws = wb["Заказы"]
        self.assertEqual(
            [cell.value for cell in ws[1]],
            [
                "Номер заказа",
                "Подготовка завершена",
                "Анализ завершён",
                "Задач обследовано",
                "Разрытие отсутствует",
                "Закрыто легально",
                "Закрыто нелегально",
                "Ещё без исхода",
            ],
        )
        self.assertEqual(ws["A2"].value, "B-2")
        self.assertEqual(ws["B2"].value, "Да")
        self.assertEqual(ws["C2"].value, "Нет")
        self.assertEqual(ws["D2"].value, 4)

    def test_nested_rows_write_parent_then_children(self) -> None:
        parent_ds = DATASETS["closed_orders"]
        child_ds = DATASETS["closed_tasks"]
        parent = SheetData(
            id="orders",
            title="Заказы с задачами",
            dataset="closed_orders",
            columns=[
                parent_ds.column("task_number"),
                parent_ds.column("rayon"),
            ],
            rows=[
                {
                    "order_key": "11111111-1111-1111-1111-111111111111",
                    "task_number": "A-1",
                    "rayon": "Арбат",
                }
            ],
        )
        child = SheetData(
            id="tasks",
            title="Задачи",
            dataset="closed_tasks",
            columns=[
                child_ds.column("order_key"),
                child_ds.column("closure_kind"),
                child_ds.column("task_key"),
            ],
            rows=[
                {
                    "order_key": "11111111-1111-1111-1111-111111111111",
                    "closure_kind": "done_legal",
                    "task_key": "task-1",
                }
            ],
        )
        parent.nested_children.append(child)
        raw = build_workbook_bytes("Вложенный", [parent])
        wb = load_workbook(BytesIO(raw))
        ws = wb["Заказы с задачами"]
        self.assertEqual(ws["A1"].value, "Тип строки")
        self.assertEqual(ws["A2"].value, "Заказ")
        self.assertEqual(ws["B2"].value, "A-1")
        self.assertEqual(ws["A3"].value, "Задача")
        self.assertEqual(ws["E3"].value, "Закрыто легально")

    def test_related_sheet_and_filename_header(self) -> None:
        dataset = DATASETS["field_summary"]
        sheets = [
            SheetData(
                id="field",
                title="Поле[]:*?",
                dataset="field_summary",
                columns=[dataset.column("user_login"), dataset.column("orders_closed")],
                rows=[{"user_login": "vasya", "orders_closed": 3}],
            )
        ]
        raw = build_workbook_bytes("Отчёт", sheets)
        wb = load_workbook(BytesIO(raw))
        self.assertEqual(wb.sheetnames[0], "Поле")
        header = content_disposition("Закрытые заказы", "2026-08-01", "2026-08-19")
        self.assertIn("filename=", header)
        self.assertIn("filename*=UTF-8''", header)

    def test_datetime_cells(self) -> None:
        dataset = DATASETS["closed_orders"]
        sheets = [
            SheetData(
                id="orders",
                title="Даты",
                dataset="closed_orders",
                columns=[dataset.column("closed_at")],
                rows=[{"closed_at": "2026-08-19T12:00:00+00:00"}],
            )
        ]
        raw = build_workbook_bytes("Даты", sheets)
        wb = load_workbook(BytesIO(raw))
        value = wb.active["A2"].value
        self.assertIsInstance(value, datetime)


class NestModeConstantTests(unittest.TestCase):
    def test_nested_constant(self) -> None:
        self.assertEqual(NEST_NESTED, "nested_rows")


class SurveyedOutcomeTests(unittest.TestCase):
    def test_priority_illegal_over_legal_and_clear(self) -> None:
        self.assertEqual(
            resolve_surveyed_task_outcome({"clear", "done_legal", "done_illegal"}),
            "done_illegal",
        )
        self.assertEqual(resolve_surveyed_task_outcome({"clear", "done_legal"}), "done_legal")
        self.assertEqual(resolve_surveyed_task_outcome({"clear", "open"}), "clear")
        self.assertEqual(resolve_surveyed_task_outcome({"open"}), "open")
        self.assertIsNone(resolve_surveyed_task_outcome([]))

    def test_aggregate_counts_by_order(self) -> None:
        snapshots = [
            {"order_key": "o1", "task_key": "t1", "closure_kind": "clear"},
            {"order_key": "o1", "task_key": "t1", "closure_kind": "done_legal"},
            {"order_key": "o1", "task_key": "t2", "closure_kind": "done_illegal"},
            {"order_key": "o1", "task_key": "t3", "closure_kind": "clear"},
            {"order_key": "o2", "task_key": "t9", "closure_kind": "done_legal"},
        ]
        active = [
            {"order_key": "o1", "task_key": "t1", "field_observed": True},
            {"order_key": "o1", "task_key": "t4", "field_observed": True},
            {"order_key": "o1", "task_key": "t5", "field_observed": False},
            {"order_key": "o2", "task_key": "t9", "field_observed": True},
        ]
        counts = aggregate_surveyed_order_counts(snapshots, active)
        self.assertEqual(
            counts["o1"],
            {
                "tasks_surveyed": 4,
                "tasks_clear": 1,
                "tasks_done_legal": 1,
                "tasks_done_illegal": 1,
                "tasks_open": 1,
            },
        )
        self.assertEqual(
            counts["o2"],
            {
                "tasks_surveyed": 1,
                "tasks_clear": 0,
                "tasks_done_legal": 1,
                "tasks_done_illegal": 0,
                "tasks_open": 0,
            },
        )


if __name__ == "__main__":
    unittest.main()
