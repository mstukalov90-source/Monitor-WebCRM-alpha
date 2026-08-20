"""Build an .xlsx workbook from report sheet data."""

from __future__ import annotations

import re
from datetime import date, datetime
from io import BytesIO
from typing import Any
from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.crm.reports.catalog import ColumnDef, format_cell_value
from app.crm.reports.query import SheetData

_UNSAFE_SHEET = re.compile(r"[\[\]:*?/\\]")
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
NESTED_FILL = PatternFill("solid", fgColor="F3F6FA")
KIND_COLUMN = ColumnDef(id="_row_kind", label="Тип строки")


def build_workbook_bytes(report_name: str, sheets: list[SheetData]) -> bytes:
    workbook = Workbook()
    workbook.properties.title = (report_name or "Отчёт")[:200]
    default = workbook.active
    first = True
    used_names: set[str] = set()

    for sheet in sheets:
        title = _unique_sheet_name(sheet.title, used_names)
        if first:
            default.title = title
            ws = default
            first = False
        else:
            ws = workbook.create_sheet(title)
        if sheet.nested_children:
            _write_nested_sheet(ws, sheet)
        else:
            _write_flat_sheet(ws, sheet.columns, sheet.rows)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def content_disposition(report_name: str, date_from: str, date_to: str) -> str:
    ascii_name = f"statistika_{date_from}_{date_to}.xlsx"
    pretty = _safe_filename(report_name or "Отчёт")
    utf_name = quote(f"{pretty}_{date_from}_{date_to}.xlsx")
    return f'attachment; filename="{ascii_name}"; filename*=UTF-8\'\'{utf_name}'


def _write_flat_sheet(
    ws: Worksheet,
    columns: list[ColumnDef],
    rows: list[dict[str, Any]],
) -> None:
    _write_header(ws, columns)
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, column in enumerate(columns, start=1):
            ws.cell(r_idx, c_idx, _excel_value(column, row.get(column.id)))
    _finish_sheet(ws, len(columns), len(rows) + 1)


def _write_nested_sheet(ws: Worksheet, sheet: SheetData) -> None:
    child_blocks: list[tuple[str, list[ColumnDef]]] = []
    for child in sheet.nested_children:
        child_cols = [col for col in child.columns if col.id not in {c.id for c in sheet.columns}]
        child_blocks.append((child.id, child_cols))

    header_cols = [KIND_COLUMN, *sheet.columns]
    for _, child_cols in child_blocks:
        header_cols.extend(child_cols)

    _write_header(ws, header_cols)

    children_by_parent: list[tuple[SheetData, dict[str, list[dict[str, Any]]]]] = []
    for child in sheet.nested_children:
        grouped: dict[str, list[dict[str, Any]]] = {}
        for row in child.rows:
            key = str(row.get("order_key") or "")
            grouped.setdefault(key, []).append(row)
        children_by_parent.append((child, grouped))

    r_idx = 2
    for parent_row in sheet.rows:
        parent_key = str(parent_row.get("order_key") or parent_row.get("object_key") or "")
        values = ["Заказ"] + [_excel_value(col, parent_row.get(col.id)) for col in sheet.columns]
        values.extend([None] * (len(header_cols) - len(values)))
        _write_row(ws, r_idx, values)
        r_idx += 1

        for child, grouped in children_by_parent:
            child_cols = [col for col in child.columns if col.id not in {c.id for c in sheet.columns}]
            for child_row in grouped.get(parent_key, []):
                values = ["Задача"] + [
                    _excel_value(col, parent_row.get(col.id)) for col in sheet.columns
                ]
                values.extend(_excel_value(col, child_row.get(col.id)) for col in child_cols)
                _write_row(ws, r_idx, values, fill=NESTED_FILL)
                r_idx += 1

    _finish_sheet(ws, len(header_cols), r_idx - 1)


def _write_header(ws: Worksheet, columns: list[ColumnDef]) -> None:
    for idx, column in enumerate(columns, start=1):
        cell = ws.cell(1, idx, column.label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"
    ws.row_dimensions[1].height = 22


def _write_row(
    ws: Worksheet,
    row_idx: int,
    values: list[Any],
    *,
    fill: PatternFill | None = None,
) -> None:
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row_idx, col_idx, value)
        if fill is not None:
            cell.fill = fill


def _finish_sheet(ws: Worksheet, col_count: int, last_row: int) -> None:
    if col_count <= 0:
        return
    last_row = max(last_row, 1)
    ws.auto_filter.ref = f"A1:{get_column_letter(col_count)}{last_row}"
    for idx in range(1, col_count + 1):
        letter = get_column_letter(idx)
        header = str(ws.cell(1, idx).value or "")
        width = min(42, max(12, len(header) + 4))
        ws.column_dimensions[letter].width = width


def _excel_value(column: ColumnDef, value: Any) -> Any:
    formatted = format_cell_value(column, value)
    if formatted is None or formatted == "":
        return None
    if column.value_type == "datetime":
        parsed = _parse_datetime(formatted)
        return parsed if parsed is not None else formatted
    if column.value_type == "int":
        try:
            return int(formatted)
        except (TypeError, ValueError):
            return formatted
    if column.value_type == "float":
        try:
            return round(float(formatted), 4)
        except (TypeError, ValueError):
            return formatted
    return formatted


def _parse_datetime(value: Any) -> datetime | date | None:
    if isinstance(value, datetime):
        return value.replace(tzinfo=None) if value.tzinfo else value
    if isinstance(value, date):
        return value
    text = str(value)
    try:
        parsed = datetime.fromisoformat(text.replace("Z", "+00:00"))
        return parsed.replace(tzinfo=None)
    except ValueError:
        return None


def _unique_sheet_name(title: str, used: set[str]) -> str:
    base = _safe_sheet_name(title)
    candidate = base
    n = 2
    while candidate.lower() in {name.lower() for name in used}:
        suffix = f"_{n}"
        candidate = f"{base[: 31 - len(suffix)]}{suffix}"
        n += 1
    used.add(candidate)
    return candidate


def _safe_sheet_name(title: str) -> str:
    cleaned = _UNSAFE_SHEET.sub(" ", (title or "").strip()) or "Лист"
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" '")
    return cleaned[:31] or "Лист"


def _safe_filename(name: str) -> str:
    cleaned = re.sub(r"[^\w\- ()а-яА-ЯёЁ]+", "_", name, flags=re.UNICODE)
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" ._") or "Отчёт"
    return cleaned[:80]
